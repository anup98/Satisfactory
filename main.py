from common import *
import os

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


FILENAME = "Satisfactory.xlsx"
REMOVELIST = [
    "Alien Carapace",
    "Alien Organs",
    "Bacon Agaric",
    "Beryl Nut",
    "Biomass",
    "Blade Runners",
    "Blue Power Slug",
    "Chainsaw",
    "Color Cartridge",
    "Fabric",
    "Factory Cart™",
    "FICSIT Coupon",
    "Flower Petals",
    "Gas Mask",
    "Gas Filter",
    "Golden Factory Cart™",
    "Hazmat Suit",
    "Hover Pack",
    "HUB Parts",
    "Iodine Infused Filter",
    "Jetpack",
    "Leaves",
    "Liquid Biofuel",
    "Medicinal Inhaler",
    "Mycelia",
    "Nobelisk Detonator",
    "Nobelisk",
    "Object Scanner",
    "Packaged Liquid Biofuel",
    "Paleberry",
    "Parachute",
    "Power Shard",
    "Purple Power Slug",
    "Portable Miner",
    "Rebar Gun",
    "Rifle Cartridge",
    "Rifle",
    "Xeno-Basher",
    "Xeno-Zapper",
    "Yellow Power Slug",
    "Zipline", 
    "Solid Biofuel", 
    "Spiked Rebar",
    "Wood",
    "Beacon"
    ]
RECIPE_PREFERENCES = {
    "Adaptive Control Unit": None,
    "AI Limiter": None,
    "Alclad Aluminum Sheet": None,
    "Alumina Solution": None,
    "Aluminum Casing": None,
    "Aluminum Ingot": None,
    "Aluminum Scrap": None,
    "Assembly Director System": None,
    "Automated Wiring": None,
    "Battery": None,
    "Bauxite": None,
    "Black Powder": None,
    "Cable": None,
    "Caterium Ingot": None,
    "Caterium Ore": None,
    "Circuit Board": None,
    "Coal": None,
    "Compacted Coal": None,
    "Computer": None,
    "Concrete": "Wet Concrete",
    "Cooling System": None,
    "Copper Ingot": None,
    "Copper Ore": None,
    "Copper Powder": None,
    "Copper Sheet": None,
    "Crude Oil": None,
    "Crystal Oscillator": None,
    "Electromagnetic Control Rod": None,
    "Empty Canister": None,
    "Empty Fluid Tank": None,
    "Encased Industrial Beam": None,
    "Encased Plutonium Cell": None,
    "Encased Uranium Cell": None,
    "Fuel": None,
    "Fused Modular Frame": None,
    "Heat Sink": None,
    "Heavy Modular Frame": None,
    "Heavy Oil Residue": None,
    "High-Speed Connector": None,
    "Iron Ingot": "Pure Iron Ingot",
    "Iron Ore": None,
    "Iron Plate": None,
    "Iron Rod": "Steel Rod",
    "Limestone": None,
    "Magnetic Field Generator": None,
    "Modular Engine": None,
    "Modular Frame": None,
    "Motor": None,
    "Nitric Acid": None,
    "Nitrogen Gas": None,
    "Non-fissile Uranium": None,
    "Nuclear Pasta": None,
    "Packaged Alumina Solution": None,
    "Packaged Fuel": None,
    "Packaged Heavy Oil Residue": None,
    "Packaged Nitric Acid": None,
    "Packaged Nitrogen Gas": None,
    "Packaged Oil": None,
    "Packaged Sulfuric Acid": None,
    "Packaged Turbofuel": None,
    "Packaged Water": None,
    "Petroleum Coke": None,
    "Plastic": None,
    "Plutonium Fuel Rod": None,
    "Plutonium Pellet": None,
    "Plutonium Waste": None,
    "Polymer Resin": None,
    "Pressure Conversion Cube": None,
    "Quartz Crystal": None,
    "Quickwire": None,
    "Radio Control Unit": None,
    "Raw Quartz": None,
    "Reinforced Iron Plate": None,
    "Rotor": None,
    "Rubber": None,
    "Screw": None,
    "Silica": None,
    "Smart Plating": None,
    "Stator": None,
    "Steel Beam": None,
    "Steel Ingot": None,
    "Steel Pipe": None,
    "Sulfur": None,
    "Sulfuric Acid": None,
    "Supercomputer": None,
    "Thermal Propulsion Rocket": None,
    "Turbo Motor": None,
    "Turbofuel": None,
    "Uranium Fuel Rod": None,
    "Uranium Waste": None,
    "Uranium": None,
    "Versatile Framework": None,
    "Water": None,
    "Wire": None
    }

def header(ws, name, row, col):
    ws.cell(row, col).value = name
    ws.cell(row, col).alignment = Alignment(horizontal='center')
    ws.merge_cells(coord(row, col, row, col + 3))

def writeInputsOutputs(ws, row, col, data: dict):
    r, c = row, col
    for i in sorted(data.keys()):
        ws.cell(r, c).value = i
        ws.cell(r, c + 1).value = data[i]
        r += 1 
    return r

def writeFormulas(ws, row, rowOffset, col, data, item):
    rowDiff = rowOffset - row
    # Desired Output
    desiredOutput = 780
    ws.cell(rowOffset, col).value = desiredOutput
    ws.cell(rowOffset, col + 1).value = "Desired Output"

    # Building
    ws.cell(rowOffset, col + 2).value = "=" + coord(rowOffset, col) + " / " + getQuantityLocation(ws, rowOffset, col + 2, item)
    ws.cell(rowOffset, col + 3).value = data.building
    rowDiff += 1

    # Total Inputs
    for inputRow in range(row, rowOffset):
        ws.cell(row + rowDiff, col).value = "=" + coord(row + rowDiff, col + 2) + " / 780"
        ws.cell(row + rowDiff, col + 1).value = "Mk 5 Lines"
        ws.cell(row + rowDiff, col + 2).value = "=" + coord(inputRow, col + 1) + " * " + coord(rowOffset, col + 2)
        ws.cell(row + rowDiff, col + 3).value = ws.cell(inputRow, col).value
        rowDiff += 1
    
    return row + rowDiff - 1



def getQuantityLocation(ws, row, col, item):
    for i in range(1, 5):
        if ws.cell(row - i, col).value == item:
            return coord(row - i, col + 1)
    return None

        


def addRows(masterItems, masterRecipes, ws):
    col = 1

    # Each Item
    for item in sorted(masterItems.keys()):
        if len(masterItems[item].recipes) <= 0:
            continue
        row = 1

        # Each Recipe
        for rName in sorted(masterItems[item].recipes, key= lambda x: (masterRecipes[x].alternate)):
            if masterRecipes[rName].alternate:
                header(ws, "Alternate: " + rName, row, col)
            else:
                header(ws, rName, row, col)

            row += 1
            
            # Each Input/Output
            inputs = writeInputsOutputs(ws, row, col, masterRecipes[rName].inputs)
            outputs = writeInputsOutputs(ws, row, col + 2, masterRecipes[rName].outputs)

            rowOffset = inputs if inputs > outputs else outputs
            row = writeFormulas(ws, row, rowOffset, col, masterRecipes[rName], item)



            # row += 1            
        col += 4

def findEmptyRow(ws, col):
    row = 2
    emptyFound = False
    while emptyFound:
        if ws.cell(row, col) == None:
            emptyFound = True
            break
        row += 1


def main():
    masterItems, masterRecipes = getItems()
    for i in REMOVELIST:
        masterItems.pop(i)

    # for i in sorted(masterItems.keys()):
    #     print(sorted(masterItems[i].recipes, key= lambda x: (masterRecipes[x].alternate)))

    # for i in sorted(masterItems.keys()):
    #     for j in masterItems[i].recipes:
    #         print(masterRecipes[j].outputs)
    
    wb = load_workbook(FILENAME)
    ws = wb.active

    addRows(masterItems, masterRecipes, ws)
        
    wb.save(FILENAME)

    
    
    
if __name__ == "__main__":
    main()
