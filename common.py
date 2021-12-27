import json
from item import Item
from recipe import Recipe
from openpyxl.utils import get_column_letter



def getItems():
    with open("data.json") as f:
        data = json.load(f)

    items = data['items']
    buildings = data['buildings']
    masterItems = {
        items[i]['name'].strip(): Item(items[i]['name'].strip()) for i in items
    }
    masterRecipes = {}

    for recipe in data['recipes'].values():
        if recipe['forBuilding'] or not recipe['inMachine']:
            continue

        r = Recipe(recipe['name'].replace("Alternate:", "").strip(), buildings[recipe['producedIn'][0]]['name'], recipe['alternate'])

        multiplier = 60 / recipe['time']

        for item in recipe['ingredients']:
            name = items[item['item']]['name']
            r.add_input(name, item['amount'] * multiplier)
            if r.name not in masterItems[name].usedInRecipes:
                masterItems[name].usedInRecipes.append(r.name)

        for item in recipe['products']:
            name = items[item['item']]['name']
            r.add_output(name, item['amount'] * multiplier)
            if r.name not in masterItems[name].recipes:
                masterItems[name].recipes.append(r.name)
        
        masterRecipes[r.name] = r
                  
    return masterItems, masterRecipes

def formulaData(r: Recipe):
    data = {
        "inputs": r.inputs,
        "outputs": r.outputs,
        "building": r.building,
    }
    return data

def coord(row1, col1, row2=None, col2=None):
    if col2 is None:
        return get_column_letter(col1) + str(row1)
    return get_column_letter(col1) + str(row1) + ":" + get_column_letter(col2) + str(row2)

